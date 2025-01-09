import os
import re
import PyPDF2
import openpyxl
from datetime import datetime

link_diarios = 'https://dejt.jt.jus.br/dejt/f/n/diariocon'
def extract_information(file_path):
    with open(file_path, 'rb') as pdf_file:
        pdf = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
    return text

def find_reclamado_info(text):
    cases = []
    regex = r"ATOrd"
    partes = re.split(regex, text)
    partes = [parte.strip() for parte in partes if parte.strip()]

    for parte in partes:
        texto_dividido = parte.split('RECLAMADO', 1)
        possui_reclamado = len(texto_dividido) == 2
        if possui_reclamado and 'ADVOGADO' not in texto_dividido[1]:
            reclamado = texto_dividido[1].split('\n')[0].replace('- ', '').strip()
            numero_processo = texto_dividido[0].split('\n')[0].replace('- ', '').strip()[:25]
            reclamante = texto_dividido[0]
            match = re.search(r'RECLAMANTE ([^\n]+)', reclamante)

            if match:
                nome = match.group(1).replace('- ', '').strip()
            else:
                nome = ''

            if reclamado and numero_processo:
                cases.append((reclamado, numero_processo, nome))

    return cases

def write_to_excel(cases, output_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plan1"

    # Adicionar os cabeçalhos
    sheet['A1'] = "Reclamado"
    sheet['B1'] = "Número do Processo"
    sheet['C1'] = "Reclamante"
    sheet['D1'] = "Orgão Julgador"
    sheet['E1'] = "Data de distribuição"
    sheet['F1'] = "Valor da Causa"
    sheet['G1'] = "Assuntos"
    sheet['H1'] = "CPF/CNPJ"

    # Adicionar os dados abaixo dos cabeçalhos
    for i, (reclamado, numero_processo, nome) in enumerate(cases, start=2):
        sheet.cell(row=i, column=1, value=reclamado)  # Coluna A: Nome do Reclamado
        sheet.cell(row=i, column=2, value=numero_processo)  # Coluna B: Número do Processo
        sheet.cell(row=i, column=3, value=nome)  # Coluna C: Nome do Reclamante
        # Coluna D reservada para CPF/CNPJ, pode ser ajustada conforme necessidade

    workbook.save(output_path)

current_date = datetime.now().strftime("%Y-%m-%d")

# Pasta onde estão os arquivos PDF
folder_path = "Trabalhista"

# Listar todos os arquivos na pasta
arquivos = os.listdir(folder_path)

# Filtrar apenas os arquivos PDF
arquivos_pdf = [arquivo for arquivo in arquivos if arquivo.endswith('.pdf')]

# Lista para armazenar todas as informações extraídas
all_cases = []

# Processar cada arquivo PDF na pasta
for pdf_file in arquivos_pdf:
    pdf_path = os.path.join(folder_path, pdf_file)
    text = extract_information(pdf_path)
    reclamado_info = find_reclamado_info(text)
    all_cases.extend(reclamado_info)

# Criar um único arquivo do Excel com todas as informações
output_excel_path = f"TRT.xlsx"
write_to_excel(all_cases, output_excel_path)
